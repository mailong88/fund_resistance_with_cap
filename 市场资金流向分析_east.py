import time
import akshare as ak
import pandas as pd
from datetime import datetime, timedelta
import myproxy
import openpyxl
from openpyxl.styles import PatternFill
import requests
import warnings
from io import BytesIO, StringIO
import vcr
import sys
import os

# 导入Mairui API
sys.path.insert(0, os.path.dirname(__file__))
from Mairui import MairuiAPI

# 初始化Mairui API客户端
mairui_api = MairuiAPI('B1C1C1CE-DD24-499D-A677-8DCCF0652730')


# def stock_summary(date):
#     """获取上交所成交额数据"""
#     params = {
#         "sqlId": "COMMON_SSE_SJ_GPSJ_CJGK_MRGK_C",
#         "PRODUCT_CODE": "17",
#         "type": "inParams",
#         "SEARCH_DATE": date,
#         "_": "1757487632413"
#     }
#
#     url = "https://query.sse.com.cn/commonQuery.do"
#     headers = {
#         "Referer": "http://www.sse.com.cn/",
#         "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
#                       "Chrome/89.0.4389.90 Safari/537.36",
#     }
#
#     filename = f'D:/cassette/{date}上证成交额.yml'
#     with vcr.use_cassette(filename):
#         r = requests.get(url, params=params, headers=headers)
#         data_json = r.json()
#     temp_df = pd.DataFrame(data_json["result"])
#     temp_df.reset_index(inplace=True)
#     return temp_df


def remove_comma(x):
    # 如果是NaN/None，直接返回原值
    if pd.isna(x):
        return x
    # 转为字符串，替换逗号
    str_x = str(x).replace(",", "")
    # 尝试转回数值（如果原本是数值型，可选步骤）
    try:
        return float(str_x) if "." in str_x else int(str_x)
    # 非数值则返回处理后的字符串
    except ValueError:
        return str_x


# def stock_szse_summary(date: str = "20240830") -> pd.DataFrame:
#     """获取深交所成交额数据"""
#     url = "http://www.szse.cn/api/report/ShowReport"
#     params = {
#         "SHOWTYPE": "xlsx",
#         "CATALOGID": "1803_sczm",
#         "TABKEY": "tab1",
#         "txtQueryDate": date,
#         "random": "0.39339437497296137",
#     }
#     filename = f'D:/cassette/{date}深证成交额.yml'
#     with vcr.use_cassette(filename):
#         r = requests.get(url, params=params)
#     with warnings.catch_warnings(record=True):
#         warnings.simplefilter("always")
#         temp_df = pd.read_excel(BytesIO(r.content), engine="openpyxl")
#     temp_df["证券类别"] = temp_df["证券类别"].str.strip()
#     temp_df.iloc[:, 2:] = temp_df.iloc[:, 2:].applymap(remove_comma)
#     temp_df.columns = ["证券类别", "数量", "成交金额", "总市值", "流通市值"]
#     temp_df["数量"] = pd.to_numeric(temp_df["数量"], errors="coerce")
#     temp_df["成交金额"] = pd.to_numeric(temp_df["成交金额"], errors="coerce")
#     temp_df["总市值"] = pd.to_numeric(temp_df["总市值"], errors="coerce")
#     temp_df["流通市值"] = pd.to_numeric(temp_df["流通市值"], errors="coerce")
#     return temp_df


def get_market_amount_range(start_date, end_date):
    """
    获取沪深两市日期范围内的总成交额（使用Mairui API）
    
    参数:
        start_date: 开始日期，格式'YYYY-MM-DD'
        end_date: 结束日期，格式'YYYY-MM-DD'
    
    返回:
        DataFrame，包含日期和市场成交额
    """
    # 转换日期格式：YYYY-MM-DD -> YYYYMMDD
    start_date_str = start_date.replace('-', '')
    end_date_str = end_date.replace('-', '')
    
    try:
        # 获取上证指数（000001.SH）的历史数据
        sh_df = mairui_api.get_index_history(
            '000001.SH', 
            start_date=start_date_str, 
            end_date=end_date_str,
            freq='d'
        )
        
        # 获取深证成指（399001.SZ）的历史数据
        sz_df = mairui_api.get_index_history(
            '399001.SZ', 
            start_date=start_date_str, 
            end_date=end_date_str,
            freq='d'
        )
        
        if sh_df.empty or sz_df.empty:
            print(f"获取市场成交额数据失败")
            return pd.DataFrame()
        
        # 重命名列
        sh_df = sh_df.rename(columns={'a': 'amount', 't': 'time'})
        sz_df = sz_df.rename(columns={'a': 'amount', 't': 'time'})
        
        # 提取日期和时间列
        sh_df['date'] = pd.to_datetime(sh_df['time']).dt.date
        sz_df['date'] = pd.to_datetime(sz_df['time']).dt.date
        
        # 合并两个指数的数据
        merged_df = pd.merge(
            sh_df[['date', 'amount']],
            sz_df[['date', 'amount']],
            on='date',
            suffixes=('_sh', '_sz'),
            how='inner'
        )
        
        # 计算总成交额（单位：亿元）
        merged_df['market_amount'] = (merged_df['amount_sh'] + merged_df['amount_sz']) / 1e8
        
        # 返回日期和市场成交额
        result_df = merged_df[['date', 'market_amount']].copy()
        result_df.columns = ['trade_date', 'market_amount']
        
        # 判断是否在交易时段内
        now = datetime.now()
        is_trading_time = False
        
        # 使用交易日历判断当前是否为交易日
        try:
            # 获取交易日历
            trade_cal = ak.tool_trade_date_hist_sina()
            today_str = now.strftime('%Y-%m-%d')
            
            # 检查今天是否为交易日
            is_trading_day = today_str in trade_cal['trade_date'].values
            
            if is_trading_day:
                # 如果今天是交易日，再判断是否在交易时段内
                hour = now.hour
                minute = now.minute
                current_time = hour * 100 + minute  # 转换为HHMM格式，如930表示9:30
                
                # 上午交易时段：9:30-11:30
                if 930 <= current_time <= 1130:
                    is_trading_time = True
                # 下午交易时段：13:00-15:00
                elif 1300 <= current_time <= 1500:
                    is_trading_time = True
        except Exception as e:
            print(f"获取交易日历失败: {e}，使用默认判断逻辑")
            # 如果获取交易日历失败，回退到简单的weekday判断
            if now.weekday() < 5:  # 0-4 表示周一至周五
                hour = now.hour
                minute = now.minute
                current_time = hour * 100 + minute
                
                # 上午交易时段：9:30-11:30
                if 930 <= current_time <= 1130:
                    is_trading_time = True
                # 下午交易时段：13:00-15:00
                elif 1300 <= current_time <= 1500:
                    is_trading_time = True
        
        # 如果在交易时段内，获取本交易日的实时数据
        if is_trading_time:
            today_str = now.strftime('%Y-%m-%d')
            today_date = now.date()
            
            # 检查历史数据中是否已包含今日数据
            has_today_data = any(pd.to_datetime(row['trade_date']).date() == today_date for _, row in result_df.iterrows())
            
            if not has_today_data:
                print(f"当前在交易时段内，正在获取本交易日({today_str})的实时市场成交额数据...")
                
                try:
                    # 获取上证指数本日分时数据
                    sh_intraday_df = mairui_api.get_index_intraday('000001.SH', freq='d')
                    
                    # 获取深证成指本日分时数据
                    sz_intraday_df = mairui_api.get_index_intraday('399001.SZ', freq='d')
                    
                    if not sh_intraday_df.empty and not sz_intraday_df.empty:
                        # 提取最新一条数据（当前最新的成交额）
                        sh_latest = sh_intraday_df.iloc[-1]
                        sz_latest = sz_intraday_df.iloc[-1]
                        
                        # 获取成交额数据
                        sh_amount = sh_latest['amount'] if 'amount' in sh_latest else sh_latest['a']
                        sz_amount = sz_latest['amount'] if 'amount' in sz_latest else sz_latest['a']
                        
                        # 计算总成交额（单位：亿元）
                        today_market_amount = (sh_amount + sz_amount) / 1e8
                        
                        # 创建本日数据行
                        today_row = pd.DataFrame({
                            'trade_date': [pd.to_datetime(today_date)],
                            'market_amount': [today_market_amount]
                        })
                        today_row['trade_date'] = pd.to_datetime(today_row['trade_date']).dt.date
                        # 合并到结果DataFrame
                        result_df = pd.concat([result_df, today_row], ignore_index=True)
                        print(f"成功获取本交易日市场成交额: {today_market_amount:.2f}亿元")
                    else:
                        print("获取本交易日实时数据失败，仅使用历史数据")
                except Exception as e:
                    print(f"获取本交易日实时数据时出错: {e}，仅使用历史数据")
        
        return result_df
        
    except Exception as e:
        print(f"获取市场成交额数据失败: {e}")
        return pd.DataFrame()


def get_market_amount(date):
    """获取沪深两市当日总成交额（使用Mairui API）"""
    # 处理 datetime 或 date 类型
    if hasattr(date, 'date'):
        # 如果是 datetime 类型，提取 date 部分
        date = date.date()
    str_date = date.strftime('%Y-%m-%d')
    
    try:
        # 使用新的方法获取日期范围内的数据
        start_date_str = str_date
        end_date_str = str_date
        
        df = get_market_amount_range(start_date_str, end_date_str)
        
        if not df.empty:
            # 返回成交额（单位：亿元）
            return df.loc[0, 'market_amount']
        else:
            return None
    except Exception as e:
        print(f"获取{str_date}市场成交额失败: {e}")
        return None


def get_sector_data(start_date, end_date, proxy_manager=None):
    """获取A股板块资金流向数据"""
    count = 0

    print("正在从东方财富获取板块列表...")
    # 东方财富获取板块列表API
    headers = {
        "host": "push2his.eastmoney.com",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36 Edg/143.0.0.0",
        "referer": "https://quote.eastmoney.com/bk/90.BK0437.html",
        "cookie": "qgqp_b_id=3bb3c3c347ff61c4c0ab9adef8144adf; st_nvi=hobeTgWpPdPmjsIKqYXEP284d; nid18=058d58319c67cfb95315e1c6cc03537f; nid18_create_time=1765161015783; gviem=lkvi9KomRUWykjdztG4X7b048; gviem_create_time=1765161015783; st_pvi=15444912710922; st_sp=2025-02-23%2000%3A54%3A57; st_inirUrl=https%3A%2F%2Fquote.eastmoney.com%2Fcenter%2Fhszs.html",
        "accept-encoding": "gzip, deflate"
    }

    # 获取板块列表
    sector_list_url = "http://push2.eastmoney.com/api/qt/clist/get"
    resp_lst = []
    page = 1

    my_vcr = vcr.VCR(
        filter_query_parameters=['_']  # 忽略动态时间戳参数_，避免请求不匹配
    )
    while 1:
        params = {
            'pn': str(page),
            'pz': '500',
            'po': '1',
            'np': '1',
            'fltt': '2',
            'invt': '2',
            'fid': 'f3',
            'fs': 'm:90+t:3',  # 概念板块
            'fields': 'f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f12,f13,f14,f15,f16,f17,f18,f20,f21,f23,f24,f25,f26,f22,f33,f11,f62,f128,f136,f115,f152',
            '_': str(int(datetime.now().timestamp() * 1000))
        }

        filename = f'./板块列表/板块列表_{page}.yml'
        if proxy_manager:
            proxy = proxy_manager.get_current_proxy()
            with my_vcr.use_cassette(filename):
                print(f"使用代理IP: {proxy['http']}")
                resp = requests.get(sector_list_url, params=params, headers=headers, proxies=proxy)
        else:
            with my_vcr.use_cassette(filename):
                resp = requests.get(sector_list_url, params=params, headers=headers)
                print("未使用代理")
        data = resp.json()
        resp_lst.extend(data['data']['diff'])
        if len(data['data']['diff']) < 100:
            break
        page += 1

    drop_lst = ['昨日', '上证', '深证', '标准', 'AH', 'HS', 'MS', '证金', '中证', '深成', '央视', 'AB', '创业', '预亏', '微盘', '百元', '纳米',
                '麒麟', '同步', '20', '红利', '价值', '周期', '转债', '中证', '基金', 'QF', '参股', 'B股', '举牌', '科创', '超级', '融资', '沪股',
                '破净', '富时', '机构', '预盈', '股权', '低价', '最近', '东方', '贬值', '长期', '微利', '小盘']
    dropped_lst = []
    for i in resp_lst:
        sector_name = i['f14'][:2]
        if sector_name not in drop_lst:
            try:
                if i['f6'] > 20e9:
                    dropped_lst.append(i)
            except Exception as e:
                print(i)
    sectors = pd.DataFrame(dropped_lst)
    sectors = sectors[['f12', 'f14', 'f6']]  # 板块代码、名称和流通值(f6)
    sectors.columns = ['code', 'name', 'market_cap']

    if sectors.empty:
        print("获取东方财富板块列表失败")
        return None
    print(f"共获取到{len(sectors)}个板块")

    # 获取交易日历(仍使用akshare)
    trade_cal = ak.tool_trade_date_hist_sina()
    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    mask = (trade_cal['trade_date'] >= start_dt) & (trade_cal['trade_date'] <= end_dt)
    trading_days = len(trade_cal[mask])

    trade_cal['trade_date'] = pd.to_datetime(trade_cal['trade_date']).dt.date

    # 3. 转换起止日期为date对象（你的原有代码）
    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

    # 4. 筛选：获取start_dt和end_dt之间的日期（包含起止日期）
    filtered_cal = trade_cal[
        (trade_cal['trade_date'] >= start_dt) &  # 大于等于起始日
        (trade_cal['trade_date'] <= end_dt)  # 小于等于结束日
        ]

    # 可选：重置索引，方便后续使用
    filtered_cal = filtered_cal.reset_index(drop=True)
    # 确保 trade_date 是 datetime64[ns] 类型
    filtered_cal['trade_date'] = pd.to_datetime(filtered_cal['trade_date'], format='%Y-%m-%d')
    
    # 获取窗口结束日期的市场成交额（使用Mairui API一次性获取）
    print(f"获取市场成交额数据 ({start_date} ~ {end_date})...")
    market_amount_df = get_market_amount_range(start_date, end_date)
    
    if market_amount_df.empty:
        print("获取市场成交额数据失败，使用备用方法...")
        filtered_cal['market_amount'] = filtered_cal['trade_date'].apply(lambda x: get_market_amount(x) * 1e8 if get_market_amount(x) else 0)
    else:
        # 确保日期类型一致
        market_amount_df['trade_date'] = pd.to_datetime(market_amount_df['trade_date'])
        # 合并市场成交额数据
        filtered_cal = pd.merge(filtered_cal, market_amount_df[['trade_date', 'market_amount']], 
                                on='trade_date', how='left')
        # 填充缺失值
        filtered_cal['market_amount'] = filtered_cal['market_amount'].fillna(0)
    # lst_market_amount = []
    # for d in range(len(mask)):
    #     window_end_date_str = filtered_cal[d].strftime('%Y-%m-%d')
    #     market_amount = get_market_amount(window_end_date_str)
    #     if market_amount is not None:
    #         lst_market_amount.append({'日期': mask[d], 'market_amount': market_amount})
    # 准备多线程共享数据
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    all_data = []
    count = 0
    count_lock = threading.Lock()
    data_lock = threading.Lock()
    failed_sectors = []
    failed_lock = threading.Lock()

    def fetch_sector_data(row):
        nonlocal count
        sector_code = row['code']
        sector_name = row['name']
        sector_market_cap = row['market_cap']  # 板块流通值

        # 东方财富板块历史数据API
        history_url = "http://push2his.eastmoney.com/api/qt/stock/kline/get"
        params = {
            'secid': f'90.{sector_code}',
            'ut': '7eea3edcaed734bea9cbfc24409ed989',
            'fields1': 'f1,f2,f3,f4,f5,f6',
            'fields2': 'f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61',
            'klt': '101',  # 日线
            'fqt': '1',
            'beg': start_date.replace('-', ''),
            'end': end_date.replace('-', ''),
            '_': str(int(datetime.now().timestamp() * 1000))
        }
        try:
            # 尝试最多5次获取数据
            max_retries = 5
            retry_count = 0
            data = None

            while retry_count < max_retries:
                proxy = proxy_manager.get_current_proxy()
                try:
                    if proxy:
                        print(f"获取数据: {sector_name},使用代理IP: {proxy['http']} (第{retry_count + 1}次尝试)")
                        resp = requests.get(history_url, params=params, headers=headers, proxies=proxy, verify=False,
                                            timeout=30)
                        pass
                    else:
                        print(f"未使用代理 (第{retry_count + 1}次尝试)")
                        resp = requests.get(history_url, params=params, headers=headers, verify=False, timeout=30)

                    data = resp.json()

                    if 'data' not in data or 'klines' not in data['data']:
                        print(f"{sector_name} 第{retry_count + 1}次尝试: 无效的API响应")
                        proxy_manager.get_current_proxy()
                        retry_count += 1
                        continue

                    # 数据获取成功，跳出循环
                    break

                except Exception as req_err:
                    print(f"{sector_name} 第{retry_count + 1}次尝试失败: {req_err}")
                    proxy_manager.get_current_proxy()
                    retry_count += 1
                    continue

            # 检查是否成功获取数据
            if data is None or 'data' not in data or 'klines' not in data['data']:
                raise ValueError(f"{sector_name} 尝试{max_retries}次后仍未成功获取数据")

            records = []
            for item in data['data']['klines']:
                parts = item.split(',')
                records.append({
                    '日期': parts[0],
                    '开盘价': float(parts[1]),
                    '收盘价': float(parts[2]),
                    '最高价': float(parts[3]),
                    '最低价': float(parts[4]),
                    '成交量': float(parts[5]),
                    '成交额': float(parts[6])
                })

            sector_df = pd.DataFrame(records)

            if not sector_df.empty:
                if len(sector_df) >= trading_days:
                    sector_df['板块'] = sector_name
                    sector_df['流通值'] = sector_market_cap  # 添加流通值字段
                    with data_lock:
                        all_data.append(sector_df)
                    with count_lock:
                        count += 1
                    return sector_name, True
                else:
                    print(f"{sector_name} 数据不足{trading_days}个交易日")
                    return sector_name, False
            else:
                print(f"{sector_name} 数据为空")
                return sector_name, False
        except Exception as e:
            print(f"获取 {sector_name} 数据失败: {e}，")
            return sector_name, False

    # 使用线程池并行获取数据
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_sector_data, row) for _, row in sectors.iterrows()]
        for future in as_completed(futures):
            sector_name, success, = future.result()
            if success:
                print(f"完成获取 {sector_name} 数据")
            else:
                with failed_lock:
                    failed_sectors.append(sector_name)

    if not all_data:
        print("所有板块数据获取失败")
        return None

    # 合并所有板块数据
    combined_df = pd.concat(all_data)

    # 打印失败的板块列表
    if failed_sectors:
        print("\n" + "=" * 50)
        print(f"数据获取失败的板块列表（共 {len(failed_sectors)} 个）:")
        print("=" * 50)
        for i, sector in enumerate(failed_sectors, 1):
            print(f"{i}. {sector}")
        print("=" * 50 + "\n")

        # 对失败的板块进行重试，最多5次，每次用不同代理IP，间隔5秒
        print("开始对失败的板块进行重试...\n")
        retry_failed_sectors = failed_sectors.copy()
        retry_success_sectors = []

        for retry_round in range(1, 6):  # 最多5轮重试
            if not retry_failed_sectors:
                break

            print(f"\n--- 第 {retry_round} 轮重试（剩余 {len(retry_failed_sectors)} 个板块）---")

            # 对每个失败板块进行重试
            current_round_failed = []
            for failed_sector_name in retry_failed_sectors:
                # 从原板块列表中找到对应的板块信息
                sector_info = sectors[sectors['name'] == failed_sector_name]
                if sector_info.empty:
                    print(f"未找到板块信息: {failed_sector_name}")
                    continue

                sector_code = sector_info.iloc[0]['code']
                sector_market_cap = sector_info.iloc[0]['market_cap']

                # 重新获取数据
                history_url = "http://push2his.eastmoney.com/api/qt/stock/kline/get"
                params = {
                    'secid': f'90.{sector_code}',
                    'ut': '7eea3edcaed734bea9cbfc24409ed989',
                    'fields1': 'f1,f2,f3,f4,f5,f6',
                    'fields2': 'f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61',
                    'klt': '101',  # 日线
                    'fqt': '1',
                    'beg': start_date.replace('-', ''),
                    'end': end_date.replace('-', ''),
                    '_': str(int(datetime.now().timestamp() * 1000))
                }

                try:
                    # 获取新的代理IP
                    proxy = proxy_manager.get_current_proxy()
                    if proxy:
                        print(f"重试 {failed_sector_name}，使用代理IP: {proxy['http']}")
                        resp = requests.get(history_url, params=params, headers=headers, proxies=proxy, verify=False,
                                            timeout=30)
                    else:
                        print(f"重试 {failed_sector_name}，未使用代理")
                        resp = requests.get(history_url, params=params, headers=headers, verify=False, timeout=30)

                    data = resp.json()

                    if 'data' in data and 'klines' in data['data']:
                        records = []
                        for item in data['data']['klines']:
                            parts = item.split(',')
                            records.append({
                                '日期': parts[0],
                                '开盘价': float(parts[1]),
                                '收盘价': float(parts[2]),
                                '最高价': float(parts[3]),
                                '最低价': float(parts[4]),
                                '成交量': float(parts[5]),
                                '成交额': float(parts[6])
                            })

                        sector_df = pd.DataFrame(records)

                        if not sector_df.empty and len(sector_df) >= trading_days:
                            sector_df['板块'] = failed_sector_name
                            sector_df['流通值'] = sector_market_cap
                            all_data.append(sector_df)
                            with count_lock:
                                count += 1
                            print(f"✓ 成功获取 {failed_sector_name} 数据")
                            retry_success_sectors.append(failed_sector_name)
                        else:
                            print(f"✗ {failed_sector_name} 数据不足或为空")
                            current_round_failed.append(failed_sector_name)
                    else:
                        print(f"✗ {failed_sector_name} 无效的API响应")
                        current_round_failed.append(failed_sector_name)

                except Exception as e:
                    print(f"✗ {failed_sector_name} 重试失败: {e}")
                    current_round_failed.append(failed_sector_name)

            # 更新下一轮需要重试的列表
            retry_failed_sectors = current_round_failed

            # 如果还有板块需要重试，等待5秒后继续
            if retry_failed_sectors and retry_round < 5:
                print(f"等待5秒后进行第 {retry_round + 1} 轮重试...")
                time.sleep(5)

        # 打印重试结果
        print("\n" + "=" * 50)
        print("重试结果汇总:")
        print("=" * 50)
        if retry_success_sectors:
            print(f"重试成功的板块（共 {len(retry_success_sectors)} 个）:")
            for i, sector in enumerate(retry_success_sectors, 1):
                print(f"  {i}. {sector}")

        if retry_failed_sectors:
            print(f"\n最终获取失败的板块（共 {len(retry_failed_sectors)} 个）:")
            for i, sector in enumerate(retry_failed_sectors, 1):
                print(f"  {i}. {sector}")
        else:
            print("\n所有板块重试成功！")
        print("=" * 50 + "\n")
    else:
        print("\n所有板块数据获取成功！\n")

    # 重新合并数据（包含重试成功的数据）
    combined_df = pd.concat(all_data)
    combined_df['日期'] = pd.to_datetime(combined_df['日期'])
    return combined_df, count, filtered_cal


def calculate_flow(sector_df, period, df_market_amount, count, weight=(0.5, 0.5)):
    """计算板块强度并排序(20日滚动窗口)"""
    if sector_df is None or sector_df.empty:
        return None
    # 获取所有交易日
    dates = sector_df['日期'].unique()
    if len(dates) < period:
        print(f"数据不足{period}个交易日，无法计算")
        return None

    # 确保数据按板块名称和日期排序
    sector_df = sector_df.sort_values(['板块', '日期'])

    # 计算涨跌幅
    sector_df['涨跌幅'] = sector_df.groupby('板块')['收盘价'].pct_change() * 100
    sector_df['涨跌幅'] = sector_df['涨跌幅'].fillna(0)
    # 成交额归一
    # 获取窗口结束日期的市场成交额
    # for i in range(len(dates) - (period - 1)):
    #     window_end_date_str = dates[i].strftime('%Y-%m-%d')
    #     market_amount = get_market_amount(window_end_date_str)
    #     if market_amount is not None:
    #         df_market_amount.append({'日期':dates[i],'market_amount':market_amount})
    # df_market_amount = pd.DataFrame(lst_market_amount)

    sector_df = pd.merge(df_market_amount, sector_df, how='right', left_on='trade_date', right_on='日期')
    sector_df['归一成交额'] = sector_df['成交额'] / sector_df['market_amount']
    # 计算成交额变化量
    sector_df['成交额变化量'] = sector_df.groupby('板块')['归一成交额'].diff()
    sector_df['成交额变化量'] = sector_df['成交额变化量'].fillna(0)

    # 准备结果DataFrame
    result_list = []

    # 滑动窗口计算(20日窗口，1日步长)
    for i in range(len(dates) - (period - 1)):
        window_dates = dates[i:i + period]
        window_df = sector_df[sector_df['日期'].isin(window_dates)]

        # 计算20日累计涨跌幅
        window_ret = window_df.groupby('板块')['涨跌幅'].sum().reset_index()
        window_ret.columns = ['板块', '20日涨跌幅']

        # 计算20日成交额变化总量
        window_vol = window_df.groupby('板块')['成交额变化量'].sum().reset_index()
        window_vol.columns = ['板块', '20日成交额变化']

        # 合并计算结果
        window_result = pd.merge(window_ret, window_vol, on='板块')

        # 获取板块流通值（使用窗口最后一日的流通值）
        window_market_cap = window_df.groupby('板块')['流通值'].last().reset_index()
        window_result = pd.merge(window_result, window_market_cap, on='板块')

        # 计算资金阻力：20日成交额变化/板块流通值/20日涨跌幅
        # 注意：当涨跌幅为0时，资金阻力设为无穷大或特殊值
        window_result['资金阻力'] = window_result.apply(
            lambda row: int(((row['20日成交额变化']) / row['流通值'] / (row['20日涨跌幅'] + 30)) * 10e7) + round(
                row['20日涨跌幅'] / 100, 2)
            if row['20日成交额变化'] >= 0 else int(((row['20日成交额变化']) / row['流通值'] / (row['20日涨跌幅'] + 30)) * 10e7) - round(
                row['20日涨跌幅'] / 100, 2),
            axis=1
        )

        # 计算排名
        window_result['涨跌幅排名'] = window_result['20日涨跌幅'].rank(ascending=False, method='min')
        window_result['资金流向排名'] = window_result['20日成交额变化'].rank(ascending=False, method='min')

        # 计算强度
        window_result['强度'] = round(((count - window_result['涨跌幅排名']) * weight[0] +
                                     (count - window_result['资金流向排名']) * weight[1]) * 100 / count, 2)

        # 添加窗口结束日期
        window_result['窗口结束日期'] = window_dates[-1]

        result_list.append(window_result)

    # 合并所有窗口结果
    result_df = pd.concat(result_list)

    # 转换为宽格式 - 每行一个板块，各窗口结束日期的强度作为列
    wide_df = result_df.pivot_table(
        index='板块',
        columns='窗口结束日期',
        values=['强度'],
        aggfunc='first'
    )

    # 扁平化多级列索引
    wide_df.columns = [f"{col[1].strftime('%Y-%m-%d %H:%M')}" for col in wide_df.columns]
    wide_df = wide_df.reset_index()

    # 转换资金阻力为宽格式
    wide_resistance = result_df.pivot_table(
        index='板块',
        columns='窗口结束日期',
        values=['资金阻力'],
        aggfunc='first'
    )

    # 扁平化多级列索引
    wide_resistance.columns = [f"{col[1].strftime('%Y-%m-%d %H:%M')}" for col in wide_resistance.columns]
    wide_resistance = wide_resistance.reset_index()
    if period == 20:
        result_df.to_excel('result_df.xlsx')
        wide_resistance.to_excel('wide_resistance.xlsx')
    return wide_df, wide_resistance


def save_data(result_3d, result_20d, resistance_3d, resistance_20d, writer):
    # 处理20日数据
    result_20d.to_excel(writer, index=False, sheet_name='20日板块强度')
    rank_20d = result_20d.copy()
    for col in rank_20d.columns:
        if col != '板块':
            rank_20d[col] = rank_20d[col].rank(ascending=False, method='min')
    rank_20d.to_excel(writer, index=False, sheet_name='20日板块强度排名')

    # 保存20日资金阻力数据
    resistance_20d.to_excel(writer, index=False, sheet_name='20日资金阻力')

    # 新增：找出所有日期中排名提升超过10且进入前20的板块
    top_improvers_list = []

    # 获取所有日期列(排除'板块'列)
    date_cols = [col for col in rank_20d.columns if col != '板块']

    for i in range(1, len(date_cols)):  # 从第二列开始，因为第一列没有前一天数据
        current_date = date_cols[i]
        prev_date = date_cols[i - 2]
        # 计算排名提升
        rank_diff = pd.DataFrame({
            '日期': current_date,
            '板块': rank_20d['板块'],
            '当前排名': rank_20d[current_date],
            '排名提升': rank_20d[prev_date] - rank_20d[current_date]
        })

        # 筛选条件：排名提升>10且当前排名<=20
        daily_improvers = rank_diff[
            (rank_diff['排名提升'] > 50) &
            (rank_diff['当前排名'] <= 50)
            ]

        if not daily_improvers.empty:
            top_improvers_list.append(daily_improvers)

    if top_improvers_list:
        # 合并所有日期的结果
        all_top_improvers = pd.concat(top_improvers_list)

        # 按日期和提升幅度排序
        all_top_improvers = all_top_improvers.sort_values(
            ['日期', '排名提升'],
            ascending=[True, False]
        )

        # 保存到新sheet页
        all_top_improvers.to_excel(
            writer,
            index=False,
            sheet_name='排名提升板块'
        )

    # 处理3日数据
    result_3d.to_excel(writer, index=False, sheet_name='3日板块强度')
    rank_3d = result_3d.copy()
    for col in rank_3d.columns:
        if col != '板块':
            rank_3d[col] = rank_3d[col].rank(ascending=False, method='min')
    rank_3d.to_excel(writer, index=False, sheet_name='3日板块强度排名')

    # 获取工作表引用
    worksheet_20d = writer.sheets['20日板块强度']
    rank_worksheet_20d = writer.sheets['20日板块强度排名']
    worksheet_3d = writer.sheets['3日板块强度']
    rank_worksheet_3d = writer.sheets['3日板块强度排名']
    worksheet_resistance_20d = writer.sheets['20日资金阻力']

    # 设置列宽为12
    for col in worksheet_20d.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        worksheet_20d.column_dimensions[col_letter].width = 12

    # 冻结第1行和前3列
    worksheet_20d.freeze_panes = 'B2'
    # 创建填充样式
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    green1_fill = PatternFill(start_color='E0F7E9', end_color='E0F7E9', fill_type='solid')
    green2_fill = PatternFill(start_color='A3E0C1', end_color='A3E0C1', fill_type='solid')
    green3_fill = PatternFill(start_color='66C999', end_color='66C999', fill_type='solid')
    green4_fill = PatternFill(start_color='2E9D6E', end_color='2E9D6E', fill_type='solid')
    red1_fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
    red2_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    red3_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
    red4_fill = PatternFill(start_color='990000', end_color='990000', fill_type='solid')

    # 为20日数据设置条件格式
    date_cols_20d = [col for col in result_20d.columns if col != '板块']
    date_cols_3d = [col for col in result_3d.columns if col != '板块']

    # 创建日期映射(3日表日期到20日表日期)
    date_map = {}
    for d in date_cols_3d:
        if d in date_cols_20d:
            date_map[d] = d
        else:
            # 找到20日表中最近的日期
            dt = pd.to_datetime(d)
            closest = min(date_cols_20d, key=lambda x: abs(pd.to_datetime(x) - dt))
            date_map[d] = closest

    for row_idx in range(len(result_20d)):
        row_data_20d = result_20d.iloc[row_idx]
        row_data_3d = result_3d.iloc[row_idx]
        for col_idx, col in enumerate(date_cols_20d):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
            cell = worksheet_20d[f"{col_letter}{row_idx + 2}"]
            if isinstance(cell.value, (int, float)):
                # 检查3日强度<10则在20日表中标记绿色
                for d3, d20 in date_map.items():
                    if d20 == col:
                        if row_data_3d[d3] >= 95:
                            cell.fill = red4_fill
                            break
                        elif row_data_3d[d3] >= 90:
                            cell.fill = red3_fill
                            break
                        elif row_data_3d[d3] >= 85:
                            cell.fill = red2_fill
                            break
                        elif row_data_3d[d3] >= 80:
                            cell.fill = red1_fill
                            break
                        if cell.value > 80:
                            #     # 检查T日20日强度>80且大于T-1日
                            #     if col_idx > 0:  # 确保不是第一列
                            #         prev_col = date_cols_20d[col_idx - 1]
                            #         prev_value = row_data_20d[prev_col]
                            #         if isinstance(prev_value, (int, float)) and cell.value > prev_value:
                            #             cell.fill = blue_fill
                            #             break

                            if row_data_3d[d3] <= 20:
                                cell.fill = green4_fill
                                break
                            elif row_data_3d[d3] <= 40:
                                cell.fill = green3_fill
                                break
                            elif row_data_3d[d3] <= 60:
                                cell.fill = green2_fill
                                break
                            elif row_data_3d[d3] < 80:
                                cell.fill = green1_fill
                                break

    # 为3日数据设置条件格式
    date_cols_3d = [col for col in result_3d.columns if col != '板块']
    for row_idx in range(len(result_3d)):
        row_data = result_3d.iloc[row_idx]
        for col_idx, col in enumerate(date_cols_3d):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
            cell = worksheet_3d[f"{col_letter}{row_idx + 2}"]
            if isinstance(cell.value, (int, float)):
                if cell.value < 10:
                    cell.fill = green_fill
                elif cell.value > 90:
                    cell.fill = red_fill

    # 为排名工作表设置条件格式(前20名标记为红色，每日提升前十标记为粉色)
    # rows_to_delete = []

    # 首先标记所有前20名的单元格为红色
    # for row_idx in range(len(rank_20d)):
    #     row_data = rank_20d.iloc[row_idx]
    #     top10_days = 0
    #     for col_idx, col in enumerate(date_cols_20d):
    #         col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
    #         cell = rank_worksheet_20d[f"{col_letter}{row_idx + 2}"]
    #         if isinstance(cell.value, (int, float)) and cell.value <= 20:
    #             cell.fill = red_fill
    #             top10_days += 1

    # # 记录进入前20天数<3的行
    # if top10_days < 3:
    #     rows_to_delete.append(row_idx + 2)  # +2因为Excel行号从1开始且第一行是标题

    # 然后处理每日提升前十的板块
    # for col_idx in range(3, len(date_cols_20d)):  # 从第二列开始比较
    #     current_date = date_cols_20d[col_idx]
    #     prev_date = date_cols_20d[col_idx-3]
    #
    #     # 计算所有板块的排名提升
    #     improvements = []
    #     for row_idx in range(len(rank_20d)):
    #         prev_rank = rank_20d.iloc[row_idx][prev_date]
    #         curr_rank = rank_20d.iloc[row_idx][current_date]
    #         if isinstance(prev_rank, (int, float)) and isinstance(curr_rank, (int, float)):
    #             improvement = prev_rank - curr_rank  # 排名提升=前一日排名-当前排名
    #             improvements.append((row_idx, improvement))
    #
    #     # 找出提升幅度最大的前10个板块
    #     improvements.sort(key=lambda x: x[1], reverse=True)
    #     top_improvers = improvements[:20]
    #
    #     # 标记这些板块的当前日期单元格为粉色
    #     col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
    #     for row_idx, _ in top_improvers:
    #         cell = rank_worksheet_20d[f"{col_letter}{row_idx + 2}"]
    #         if isinstance(cell.value, (int, float)):
    #             cell.fill = pink_fill

    #     # 记录进入前10天数<3的行
    #     if top10_days < 3:
    #         rows_to_delete.append(row_idx + 2)  # +2因为Excel行号从1开始且第一行是标题

    # # 从后往前删除行，避免索引变化问题
    # for row_idx in sorted(rows_to_delete, reverse=True):
    #     rank_worksheet_20d.delete_rows(row_idx)

    # 为20日排名表添加颜色标记
    for row_idx in range(len(rank_20d)):
        row_data_20d = rank_20d.iloc[row_idx]
        row_data_3d = result_3d.iloc[row_idx]
        for col_idx, col in enumerate(date_cols_20d):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
            cell = rank_worksheet_20d[f"{col_letter}{row_idx + 2}"]
            if isinstance(cell.value, (int, float)):
                # 1. 检查是否在20日排名前10
                if cell.value <= 20:
                    cell.fill = pink_fill
                    # 2. 检查对应的3日排名是否也在前10
                    # 找到对应的3日日期
                    dt_20d = pd.to_datetime(col)
                    closest_3d_date = min(date_cols_3d, key=lambda x: abs(pd.to_datetime(x) - dt_20d))
                    # if row_data_3d[closest_3d_date] >= 90:
                    #     cell.fill = red_fill
                    if row_data_3d[closest_3d_date] >= 95:
                        cell.fill = red4_fill
                        continue
                    elif row_data_3d[closest_3d_date] >= 90:
                        cell.fill = red3_fill
                        continue
                    elif row_data_3d[closest_3d_date] >= 85:
                        cell.fill = red2_fill
                        continue
                    elif row_data_3d[closest_3d_date] >= 80:
                        cell.fill = red1_fill
                        continue
                    elif row_data_3d[closest_3d_date] <= 20:
                        cell.fill = green4_fill
                        continue
                    elif row_data_3d[closest_3d_date] <= 40:
                        cell.fill = green3_fill
                        continue
                    elif row_data_3d[closest_3d_date] <= 60:
                        cell.fill = green2_fill
                        continue
                    elif row_data_3d[closest_3d_date] < 80:
                        cell.fill = green1_fill
                        continue
    rank_worksheet_20d.freeze_panes = 'B2'
    for row_idx in range(len(rank_3d)):
        row_data = rank_3d.iloc[row_idx]
        for col_idx, col in enumerate(date_cols_3d):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
            cell = rank_worksheet_3d[f"{col_letter}{row_idx + 2}"]
            if isinstance(cell.value, (int, float)) and cell.value <= 10:
                cell.fill = red_fill

    # 为20日资金阻力工作表设置条件格式
    resistance_date_cols = [col for col in resistance_20d.columns if col != '板块']
    for row_idx in range(len(resistance_20d)):
        row_data = resistance_20d.iloc[row_idx]
        for col_idx, col in enumerate(resistance_date_cols):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 2)
            cell = worksheet_resistance_20d[f"{col_letter}{row_idx + 2}"]
            if isinstance(cell.value, (int, float)):
                if cell.value < 0:
                    cell.fill = green_fill
                elif cell.value > 50:
                    cell.fill = red_fill


def main():
    proxy_manager = myproxy.ProxyManager(num=15)
    # 获取足够长的历史数据(覆盖20日分析需求)
    # now = datetime.now() - timedelta(days=50)
    now = datetime.now()
    end_date = now.strftime('%Y-%m-%d')

    # end_date = '2026-3-2'
    d_end_date = datetime.strptime(end_date, '%Y-%m-%d')

    if now.hour < 9:
        end_date = (d_end_date - timedelta(days=1)).strftime('%Y-%m-%d')
    start_date = (d_end_date - timedelta(days=50)).strftime('%Y-%m-%d')

    print("\n正在获取A股板块资金流向数据...")

    sector_df, count, lst_market_amount = get_sector_data(start_date, end_date, proxy_manager)

    if sector_df is not None:
        # 计算20日和3日分析结果
        result_20d, resistance_20d = calculate_flow(sector_df, 10, lst_market_amount, count)
        result_3d, resistance_3d = calculate_flow(sector_df, 1, lst_market_amount, count)

        if result_20d is not None and result_3d is not None:
            try:
                output_file = f"./板块资金流向/市场资金流向east_{start_date}_{end_date}.xlsx"
                writer = pd.ExcelWriter(output_file, engine='openpyxl')
                save_data(result_3d, result_20d, resistance_3d, resistance_20d, writer)
                writer.close()
                print(f"分析完成，结果已保存到: {output_file}")
            except Exception as e:
                print(f"保存Excel文件时出错: {e}")


def get_sector_data_15min(start_date, end_date, proxy_manager):
    """获取A股板块15分钟K线数据"""
    # 复制get_sector_data函数，修改klt参数为15
    count = 0
    print("正在从东方财富获取板块列表...")
    import requests
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36'
    }

    # 获取板块列表(与get_sector_data相同)
    sector_list_url = "http://push2.eastmoney.com/api/qt/clist/get"
    resp_lst = []
    page = 1

    while 1:
        params = {
            'pn': str(page),
            'pz': '500',
            'po': '1',
            'np': '1',
            'fltt': '2',
            'invt': '2',
            'fid': 'f3',
            'fs': 'm:90+t:3',
            'fields': 'f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f12,f13,f14,f15,f16,f17,f18,f20,f21,f23,f24,f25,f26,f22,f33,f11,f62,f128,f136,f115,f152',
            '_': str(int(datetime.now().timestamp() * 1000))
        }

        try:
            proxy = proxy_manager.get_current_proxy()
            if proxy:
                resp = requests.get(sector_list_url, params=params, headers=headers, proxies=proxy, verify=False)
                print(f"使用代理IP: {proxy['http']}")
            else:
                resp = requests.get(sector_list_url, params=params, headers=headers, verify=False)
                print("未使用代理")
            data = resp.json()
            resp_lst.extend(data['data']['diff'])
            if len(data['data']['diff']) < 100:
                break
            page += 1
        except Exception as e:
            print(f"获取板块列表时发生错误: {e}")
            return None

    drop_lst = ['昨日', '上证', '深证', '标准', 'AH', 'HS', 'MS', '证金', '中证', '深成', '央视', 'AB', '破净', '百元']
    need_lst = []
    for i in resp_lst:
        sector_name = i['f14']
        if sector_name[:2] in drop_lst:
            pass
        else:
            need_lst.append(i)
    sectors = pd.DataFrame(need_lst)
    sectors = sectors[['f12', 'f14', 'f6']]
    sectors.columns = ['code', 'name', 'market_cap']

    if sectors.empty:
        print("获取东方财富板块列表失败")
        return None
    print(f"共获取到{len(sectors)}个板块")

    # 准备多线程共享数据
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    all_data = []
    count = 0
    count_lock = threading.Lock()
    data_lock = threading.Lock()

    def fetch_sector_data(row):
        nonlocal count
        sector_code = row['code']
        sector_name = row['name']
        sector_market_cap = row['market_cap']

        # 修改klt参数为15(15分钟K线)
        history_url = "http://push2his.eastmoney.com/api/qt/stock/kline/get"
        params = {
            'secid': f'90.{sector_code}',
            'ut': '7eea3edcaed734bea9cbfc24409ed989',
            'fields1': 'f1,f2,f3,f4,f5,f6',
            'fields2': 'f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61',
            'klt': '15',  # 15分钟K线
            'fqt': '1',
            'beg': start_date.replace('-', ''),
            'end': end_date.replace('-', ''),
            '_': str(int(datetime.now().timestamp() * 1000))
        }
        proxy = proxy_manager.get_current_proxy()
        try:
            # 尝试最多5次获取数据
            max_retries = 5
            retry_count = 0
            data = None

            while retry_count < max_retries:
                proxy = proxy_manager.get_current_proxy()
                try:
                    if proxy:
                        resp = requests.get(history_url, params=params, headers=headers, proxies=proxy, verify=False,
                                            timeout=30)
                        print(f"获取数据: {sector_name},使用代理IP: {proxy['http']} (第{retry_count + 1}次尝试)")
                    else:
                        resp = requests.get(history_url, params=params, headers=headers, verify=False, timeout=30)
                        print(f"未使用代理 (第{retry_count + 1}次尝试)")

                    data = resp.json()

                    if 'data' not in data or 'klines' not in data['data']:
                        print(f"{sector_name} 第{retry_count + 1}次尝试: 无效的API响应")
                        proxy_manager.switch_proxy()
                        retry_count += 1
                        continue

                    # 数据获取成功，跳出循环
                    break

                except Exception as req_err:
                    print(f"{sector_name} 第{retry_count + 1}次尝试失败: {req_err}")
                    proxy_manager.switch_proxy()
                    retry_count += 1
                    continue

            # 检查是否成功获取数据
            if data is None or 'data' not in data or 'klines' not in data['data']:
                raise ValueError(f"{sector_name} 尝试{max_retries}次后仍未成功获取数据")

            records = []
            for item in data['data']['klines']:
                parts = item.split(',')
                records.append({
                    '日期': parts[0],  # 修改列名为时间
                    '开盘价': float(parts[1]),
                    '收盘价': float(parts[2]),
                    '成交额': float(parts[6])
                })

            sector_df = pd.DataFrame(records)
            if not sector_df.empty:
                sector_df['板块'] = sector_name
                sector_df['流通值'] = sector_market_cap
                with data_lock:
                    all_data.append(sector_df)
                with count_lock:
                    count += 1
                return sector_name, True
            else:
                print(f"{sector_name} 数据为空")
                return sector_name, False
        except Exception as e:
            print(f"获取 {sector_name} 数据失败: {e}")
            return sector_name, False

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(fetch_sector_data, row) for _, row in sectors.iterrows()]
        for future in as_completed(futures):
            result = future.result()
            if result is None:
                continue
            sector_name, success = result
            if success:
                print(f"完成获取 {sector_name} 数据")

    if not all_data:
        print("所有板块数据获取失败")
        return None

    combined_df = pd.concat(all_data)
    combined_df['日期'] = pd.to_datetime(combined_df['日期'])  # 修改列名为日期
    return combined_df, count


def main_15min():
    """获取15分钟K线数据的主函数"""
    proxy_manager = myproxy.ProxyManager()
    # 获取最近3天的15分钟K线数据
    now = datetime.now()
    end_date = now.strftime('%Y-%m-%d')
    if now.hour < 9:
        end_date = (now - timedelta(days=1)).strftime('%Y-%m-%d')
    start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

    print("\n正在获取A股板块15分钟K线数据...")
    sector_df, count = get_sector_data_15min(start_date, end_date, proxy_manager)

    if sector_df is not None:
        # 计算20周期和5周期分析结果
        result_20p, resistance_20p = calculate_flow(sector_df, 48, count, weight=[0.4, 0.6])
        result_5p, resistance_5p = calculate_flow(sector_df, 3, count, weight=[0.2, 0.8])
        if result_20p is not None and result_5p is not None:
            try:
                output_file = f"./板块资金流向/市场资金15分钟流向east_{start_date}_{end_date}.xlsx"
                writer = pd.ExcelWriter(output_file, engine='openpyxl')
                save_data(result_5p, result_20p, resistance_5p, resistance_20p, writer)
                writer.close()
                print(f"分析完成，结果已保存到: {output_file}")
            except Exception as e:
                print(f"保存Excel文件时出错: {e}")


if __name__ == '__main__':
    main()
